from django.shortcuts import render
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from django.http.response import HttpResponse

# Create your views here.
from .models import Book
from .serializers import BookSerializer
from rest_framework import viewsets
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated

class BookViewSet(viewsets.ModelViewSet):
  queryset = Book.objects.all()
  serializer_class = BookSerializer

class BooksExportExcel(TemplateView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        #Books
        books = Book.objects.all()
        
        #Libro de trabajo
        wb = Workbook()
        #Hoja activa, por defecto la primera del libro
        ws = wb.active
        
        ws['B1'] = 'REPORTE DE LIBROS'
        ws.merge_cells('B1:E1')
        #Encabezados de campos
        ws['B3'] = 'ID'
        ws['C3'] = 'TÍTULO'
        ws['D3'] = 'AÑO'
        ws['E3'] = 'AUTOR'
        cont=4
        #Listado del conjunto de autores
        for book in books:
            ws.cell(row=cont,column=2).value = book.id
            ws.cell(row=cont,column=3).value = book.title
            ws.cell(row=cont,column=4).value = book.year
            ws.cell(row=cont,column=5).value = book.author.name
            cont = cont + 1
        
        nombre_archivo ="BooksExportExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
