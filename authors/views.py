from django.shortcuts import render
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from django.http.response import HttpResponse

# Create your views here.
from .models import Author
from .serializers import AuthorSerializer
from rest_framework import viewsets
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated

class AuthorViewSet(viewsets.ModelViewSet):
  queryset = Author.objects.all()
  serializer_class = AuthorSerializer

class AuthorsExportExcel(TemplateView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        #Authors
        authors = Author.objects.all()
        
        #Libro de trabajo
        wb = Workbook()
        #Hoja activa, por defecto la primera del libro
        ws = wb.active
        
        ws['B1'] = 'REPORTE DE AUTORES'
        ws.merge_cells('B1:E1')
        #Encabezados de campos
        ws['B3'] = 'ID'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'NACIMIENTO'
        cont=4
        #Listado del conjunto de autores
        for author in authors:
            ws.cell(row=cont,column=2).value = author.id
            ws.cell(row=cont,column=3).value = author.name
            ws.cell(row=cont,column=4).value = author.birthday
            cont = cont + 1
        
        nombre_archivo ="AuthorsExportExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
